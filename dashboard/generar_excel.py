"""
Genera un reporte Excel completo desde la BD puestos_2026.db
con multiples hojas, formatos, colores y graficas en cada hoja.
Salida: dashboard/reporte_congreso_boyaca_2026.xlsx
"""

import os
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import BarChart, PieChart, Reference, BarChart3D
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB_PATH = os.path.join(PROJECT_ROOT, "db", "puestos_2026.db")
OUTPUT = os.path.join(PROJECT_ROOT, "dashboard", "reporte_congreso_boyaca_2026.xlsx")

AMBS = ["0700001", "0700079", "0700277", "0700181"]

# Estilos
DARK_BG = PatternFill("solid", fgColor="1A1A2E")
HEADER_BG = PatternFill("solid", fgColor="0F3460")
ACCENT_BG = PatternFill("solid", fgColor="16213E")
GREEN_BG = PatternFill("solid", fgColor="007C34")
WHITE_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(color="FFFFFF", bold=True, size=14)
SUBTITLE_FONT = Font(color="8899AA", size=10)
DATA_FONT = Font(color="E0E6ED", size=10)
NUMBER_FONT = Font(color="E0E6ED", size=10, name="Consolas")
LINK_FONT = Font(color="2196F3", size=10, underline="single")
THIN_BORDER = Border(
    bottom=Side(style="thin", color="233345"),
    top=Side(style="thin", color="233345"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

COLORES_PARTIDOS = {
    "Alianza Verde": "007C34",
    "Pacto Historico": "7B2D8B",
    "Centro Democratico": "1E477D",
    "Conservador": "E07B00",
}

CHART_COLORS = ["2196F3", "007C34", "E07B00", "7B2D8B", "FF5252",
                "00BCD4", "FFC107", "9C27B0", "4CAF50", "FF9800"]


def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_BG
        cell.font = WHITE_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_data_row(ws, row, ncols, alt=False):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = ACCENT_BG if alt else DARK_BG
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        if isinstance(cell.value, (int, float)):
            cell.font = NUMBER_FONT
            cell.alignment = RIGHT


def auto_width(ws, ncols, min_w=10, max_w=40):
    for c in range(1, ncols + 1):
        max_len = min_w
        for row in ws.iter_rows(min_col=c, max_col=c):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)) + 2, max_w))
        ws.column_dimensions[get_column_letter(c)].width = max_len


def add_title(ws, title, subtitle, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = TITLE_FONT
    cell.fill = DARK_BG
    cell.alignment = LEFT

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    cell = ws.cell(row=2, column=1, value=subtitle)
    cell.font = SUBTITLE_FONT
    cell.fill = DARK_BG

    for r in range(1, 3):
        for c in range(1, ncols + 1):
            ws.cell(row=r, column=c).fill = DARK_BG


def fill_dark_bg(ws, max_row, max_col):
    """Rellenar celdas vacias con fondo oscuro."""
    for r in range(1, max_row + 1):
        for cc in range(1, max_col + 1):
            cell = ws.cell(row=r, column=cc)
            if cell.fill == PatternFill():
                cell.fill = DARK_BG


def sheet_resumen(wb, conn):
    ws = wb.active
    ws.title = "Resumen"
    ws.sheet_properties.tabColor = "2196F3"

    c = conn.cursor()

    c.execute("SELECT COUNT(*) FROM municipios")
    n_munis = c.fetchone()[0]
    c.execute("SELECT corporacion, COUNT(*), SUM(votos) FROM votos GROUP BY corporacion")
    totales = {r[0]: {"filas": r[1], "votos": r[2]} for r in c.fetchall()}
    c.execute("SELECT COUNT(DISTINCT zona) FROM votos")
    n_zonas = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM candidatos")
    n_cands = c.fetchone()[0]
    c.execute("SELECT COUNT(DISTINCT codpar || '-' || corporacion) FROM partidos")
    n_partidos = c.fetchone()[0]

    ncols = 8
    add_title(ws, "REPORTE ELECTORAL - CONGRESO BOYACA 2026",
              "Preconteo 8 de marzo de 2026 | Departamento de Boyaca | 4 municipios principales", ncols)

    # KPIs
    kpis = [
        ("Municipios en BD", n_munis),
        ("Zonas electorales", n_zonas),
        ("Candidatos registrados", n_cands),
        ("Partidos", n_partidos),
        ("Votos Camara (CA)", totales.get("CA", {}).get("votos", 0)),
        ("Votos Senado (SE)", totales.get("SE", {}).get("votos", 0)),
        ("Filas de votos CA", totales.get("CA", {}).get("filas", 0)),
        ("Filas de votos SE", totales.get("SE", {}).get("filas", 0)),
    ]

    row = 4
    ws.cell(row=row, column=1, value="INDICADORES GENERALES").font = WHITE_FONT
    ws.cell(row=row, column=1).fill = DARK_BG
    row += 1
    headers = ["Indicador", "Valor"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=row, column=ci, value=h)
    style_header_row(ws, row, 2)
    row += 1

    for label, val in kpis:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=val)
        style_data_row(ws, row, 2, alt=(row % 2 == 0))
        row += 1

    # --- Votos por municipio ---
    row += 1
    ws.cell(row=row, column=1, value="VOTOS POR MUNICIPIO").font = WHITE_FONT
    ws.cell(row=row, column=1).fill = DARK_BG
    row += 1
    headers = ["Municipio", "Codigo", "Votos CA", "Votos SE", "Total", "Diferencia SE-CA", "% Diferencia"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=row, column=ci, value=h)
    style_header_row(ws, row, len(headers))
    row += 1

    c.execute("""
        SELECT m.nombre, m.amb,
            SUM(CASE WHEN v.corporacion='CA' THEN v.votos ELSE 0 END) as vca,
            SUM(CASE WHEN v.corporacion='SE' THEN v.votos ELSE 0 END) as vse
        FROM municipios m
        JOIN votos v ON m.amb = v.amb
        WHERE m.amb IN (?,?,?,?)
        GROUP BY m.amb ORDER BY vca DESC
    """, AMBS)
    chart_start = row
    muni_rows = c.fetchall()
    for r in muni_rows:
        ws.cell(row=row, column=1, value=r[0])
        ws.cell(row=row, column=2, value=r[1])
        ws.cell(row=row, column=3, value=r[2])
        ws.cell(row=row, column=4, value=r[3])
        ws.cell(row=row, column=5, value=r[2] + r[3])
        ws.cell(row=row, column=6, value=r[3] - r[2])
        ws.cell(row=row, column=7, value=(r[3] - r[2]) / r[2] if r[2] > 0 else 0)
        ws.cell(row=row, column=7).number_format = '0.00%'
        style_data_row(ws, row, 7, alt=(row % 2 == 0))
        row += 1
    chart_end = row - 1

    # Grafica 1: Barras agrupadas CA vs SE por municipio
    chart = BarChart()
    chart.type = "col"
    chart.title = "Votos CA vs SE por municipio"
    chart.y_axis.title = "Votos"
    chart.x_axis.title = "Municipio"
    chart.style = 10
    chart.width = 22
    chart.height = 14
    cats = Reference(ws, min_col=1, min_row=chart_start, max_row=chart_end)
    data_ca = Reference(ws, min_col=3, min_row=chart_start - 1, max_row=chart_end)
    data_se = Reference(ws, min_col=4, min_row=chart_start - 1, max_row=chart_end)
    chart.add_data(data_ca, titles_from_data=True)
    chart.add_data(data_se, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = "2196F3"
    chart.series[1].graphicalProperties.solidFill = "007C34"
    ws.add_chart(chart, f"A{row + 1}")

    # Grafica 2: Torta de votos totales por municipio
    pie = PieChart()
    pie.title = "Distribucion de votos totales por municipio"
    pie.style = 10
    pie.width = 18
    pie.height = 14
    pie_labels = Reference(ws, min_col=1, min_row=chart_start, max_row=chart_end)
    pie_data = Reference(ws, min_col=5, min_row=chart_start - 1, max_row=chart_end)
    pie.add_data(pie_data, titles_from_data=True)
    pie.set_categories(pie_labels)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showCatName = True
    for i, color in enumerate(["2196F3", "007C34", "E07B00", "7B2D8B"]):
        if i < len(pie.series[0].data_points if hasattr(pie.series[0], 'data_points') else []):
            pass
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = color
        pie.series[0].data_points.append(pt)
    ws.add_chart(pie, f"E{row + 1}")

    # --- Participacion por zonas ---
    row += 17
    ws.cell(row=row, column=1, value="PARTICIPACION POR ZONAS").font = WHITE_FONT
    ws.cell(row=row, column=1).fill = DARK_BG
    row += 1
    headers2 = ["Municipio", "Zona", "Votos CA", "Votos SE", "Total", "Ratio SE/CA"]
    for ci, h in enumerate(headers2, 1):
        ws.cell(row=row, column=ci, value=h)
    style_header_row(ws, row, len(headers2))
    row += 1

    c.execute("""
        SELECT m.nombre, v.zona,
            SUM(CASE WHEN v.corporacion='CA' THEN v.votos ELSE 0 END) as vca,
            SUM(CASE WHEN v.corporacion='SE' THEN v.votos ELSE 0 END) as vse
        FROM municipios m
        JOIN votos v ON m.amb = v.amb
        WHERE m.amb IN (?,?,?,?)
        GROUP BY m.amb, v.zona ORDER BY m.nombre, v.zona
    """, AMBS)

    zonas_start = row
    for r in c.fetchall():
        ws.cell(row=row, column=1, value=r[0])
        ws.cell(row=row, column=2, value=r[1])
        ws.cell(row=row, column=3, value=r[2])
        ws.cell(row=row, column=4, value=r[3])
        ws.cell(row=row, column=5, value=r[2] + r[3])
        ws.cell(row=row, column=6, value=r[3] / r[2] if r[2] > 0 else 0)
        ws.cell(row=row, column=6).number_format = '0.000'
        style_data_row(ws, row, 6, alt=(row % 2 == 0))
        row += 1
    zonas_end = row - 1

    # Grafica 3: Barras apiladas CA+SE por zona
    chart3 = BarChart()
    chart3.type = "col"
    chart3.grouping = "stacked"
    chart3.title = "Votos CA + SE por zona electoral"
    chart3.y_axis.title = "Votos"
    chart3.style = 10
    chart3.width = 30
    chart3.height = 14
    cats3 = Reference(ws, min_col=2, min_row=zonas_start, max_row=zonas_end)
    d_ca = Reference(ws, min_col=3, min_row=zonas_start - 1, max_row=zonas_end)
    d_se = Reference(ws, min_col=4, min_row=zonas_start - 1, max_row=zonas_end)
    chart3.add_data(d_ca, titles_from_data=True)
    chart3.add_data(d_se, titles_from_data=True)
    chart3.set_categories(cats3)
    chart3.series[0].graphicalProperties.solidFill = "2196F3"
    chart3.series[1].graphicalProperties.solidFill = "007C34"
    ws.add_chart(chart3, f"A{row + 1}")

    auto_width(ws, ncols)
    fill_dark_bg(ws, row + 20, ncols)


def sheet_candidatos_ca(wb, conn):
    ws = wb.create_sheet("Top Candidatos CA")
    ws.sheet_properties.tabColor = "E07B00"
    ncols = 8

    add_title(ws, "TOP CANDIDATOS - CAMARA DE REPRESENTANTES",
              "Ranking por votos en cada municipio | Preconteo Congreso 2026", ncols)

    headers = ["Municipio", "Ranking", "Candidato", "Partido", "Votos",
               "% del Municipio", "% del Partido", "Cedula"]
    row = 4
    for ci, h in enumerate(headers, 1):
        ws.cell(row=row, column=ci, value=h)
    style_header_row(ws, row, ncols)
    row += 1

    c = conn.cursor()

    # Recopilar top 5 global para la grafica
    global_top = []

    for amb in AMBS:
        c.execute("SELECT SUM(votos) FROM votos WHERE corporacion='CA' AND amb=?", (amb,))
        total_muni = c.fetchone()[0] or 1

        c.execute("""
            SELECT m.nombre, ca.nomcan || ' ' || ca.apecan as candidato,
                   COALESCE(p.nombre, 'Partido ' || ca.codpar) as partido,
                   SUM(v.votos) as votos, ca.cedula, ca.codpar
            FROM votos v
            JOIN candidatos ca ON v.codcan = ca.codcan AND v.codpar = ca.codpar
                AND v.corporacion = ca.corporacion AND v.amb = ca.amb
            JOIN municipios m ON v.amb = m.amb
            LEFT JOIN partidos p ON v.codpar = p.codpar AND v.corporacion = p.corporacion
            WHERE v.corporacion = 'CA' AND v.amb = ? AND v.codcan != '0'
            GROUP BY ca.codcan, ca.codpar
            ORDER BY votos DESC
            LIMIT 15
        """, (amb,))

        ranking = 0
        for r in c.fetchall():
            ranking += 1
            c2 = conn.cursor()
            c2.execute("SELECT SUM(votos) FROM votos WHERE corporacion='CA' AND amb=? AND codpar=?",
                       (amb, r[5]))
            total_partido = c2.fetchone()[0] or 1

            ws.cell(row=row, column=1, value=r[0])
            ws.cell(row=row, column=2, value=ranking)
            ws.cell(row=row, column=3, value=r[1])
            ws.cell(row=row, column=4, value=r[2])
            ws.cell(row=row, column=5, value=r[3])
            ws.cell(row=row, column=6, value=r[3] / total_muni)
            ws.cell(row=row, column=6).number_format = '0.00%'
            ws.cell(row=row, column=7, value=r[3] / total_partido)
            ws.cell(row=row, column=7).number_format = '0.00%'
            ws.cell(row=row, column=8, value=r[4] or "")
            style_data_row(ws, row, ncols, alt=(row % 2 == 0))

            if ranking <= 3:
                global_top.append((f"{r[1]} ({r[0][:3]})", r[3], r[2]))
            row += 1

    # --- Grafica 1: Top 10 candidatos global (barras horizontales) ---
    global_top.sort(key=lambda x: x[1], reverse=True)
    top10 = global_top[:10]

    chart_row = row + 2
    ws.cell(row=chart_row, column=1, value="Candidato").font = WHITE_FONT
    ws.cell(row=chart_row, column=1).fill = HEADER_BG
    ws.cell(row=chart_row, column=2, value="Votos").font = WHITE_FONT
    ws.cell(row=chart_row, column=2).fill = HEADER_BG
    for i, (nombre, votos, partido) in enumerate(top10):
        ws.cell(row=chart_row + 1 + i, column=1, value=nombre)
        ws.cell(row=chart_row + 1 + i, column=2, value=votos)
        style_data_row(ws, chart_row + 1 + i, 2)

    chart = BarChart()
    chart.type = "bar"  # horizontal
    chart.title = "Top 10 candidatos Camara (4 municipios)"
    chart.x_axis.title = "Votos"
    chart.style = 10
    chart.width = 28
    chart.height = 16
    chart.legend = None
    cats = Reference(ws, min_col=1, min_row=chart_row + 1, max_row=chart_row + len(top10))
    data = Reference(ws, min_col=2, min_row=chart_row, max_row=chart_row + len(top10))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = "2196F3"
    ws.add_chart(chart, f"D{chart_row}")

    # --- Grafica 2: Barras por municipio (top 1 candidato de cada uno) ---
    row2 = chart_row + len(top10) + 2
    c.execute("""
        SELECT m.nombre,
               ca.nomcan || ' ' || ca.apecan as candidato,
               COALESCE(p.nombre, 'Partido ' || ca.codpar) as partido,
               SUM(v.votos) as votos
        FROM votos v
        JOIN candidatos ca ON v.codcan = ca.codcan AND v.codpar = ca.codpar
            AND v.corporacion = ca.corporacion AND v.amb = ca.amb
        JOIN municipios m ON v.amb = m.amb
        LEFT JOIN partidos p ON v.codpar = p.codpar AND v.corporacion = p.corporacion
        WHERE v.corporacion = 'CA' AND v.codcan != '0' AND v.amb IN (?,?,?,?)
        GROUP BY v.amb, ca.codcan, ca.codpar
        ORDER BY v.amb, votos DESC
    """, AMBS)

    # Agrupar por municipio, tomar top 3 de cada uno
    from collections import defaultdict
    muni_top3 = defaultdict(list)
    for r in c.fetchall():
        if len(muni_top3[r[0]]) < 3:
            muni_top3[r[0]].append((r[1], r[3]))

    ws.cell(row=row2, column=1, value="Municipio").font = WHITE_FONT
    ws.cell(row=row2, column=1).fill = HEADER_BG
    ws.cell(row=row2, column=2, value="1er lugar").font = WHITE_FONT
    ws.cell(row=row2, column=2).fill = HEADER_BG
    ws.cell(row=row2, column=3, value="2do lugar").font = WHITE_FONT
    ws.cell(row=row2, column=3).fill = HEADER_BG
    ws.cell(row=row2, column=4, value="3er lugar").font = WHITE_FONT
    ws.cell(row=row2, column=4).fill = HEADER_BG

    idx = 0
    for muni_name in ["TUNJA", "DUITAMA", "SOGAMOSO", "PAIPA"]:
        tops = muni_top3.get(muni_name, [])
        ws.cell(row=row2 + 1 + idx, column=1, value=muni_name)
        for j, (cand, votos) in enumerate(tops):
            ws.cell(row=row2 + 1 + idx, column=2 + j, value=votos)
        style_data_row(ws, row2 + 1 + idx, 4)
        idx += 1

    chart2 = BarChart()
    chart2.type = "col"
    chart2.grouping = "clustered"
    chart2.title = "Top 3 candidatos por municipio (votos)"
    chart2.y_axis.title = "Votos"
    chart2.style = 10
    chart2.width = 22
    chart2.height = 14
    cats2 = Reference(ws, min_col=1, min_row=row2 + 1, max_row=row2 + 4)
    d1 = Reference(ws, min_col=2, min_row=row2, max_row=row2 + 4)
    d2 = Reference(ws, min_col=3, min_row=row2, max_row=row2 + 4)
    d3 = Reference(ws, min_col=4, min_row=row2, max_row=row2 + 4)
    chart2.add_data(d1, titles_from_data=True)
    chart2.add_data(d2, titles_from_data=True)
    chart2.add_data(d3, titles_from_data=True)
    chart2.set_categories(cats2)
    chart2.series[0].graphicalProperties.solidFill = "FFD700"
    chart2.series[1].graphicalProperties.solidFill = "C0C0C0"
    chart2.series[2].graphicalProperties.solidFill = "CD7F32"
    ws.add_chart(chart2, f"A{row2 + 6}")

    auto_width(ws, ncols)


def sheet_partidos_se(wb, conn):
    ws = wb.create_sheet("Partidos Senado")
    ws.sheet_properties.tabColor = "7B2D8B"
    ncols = 6

    add_title(ws, "DISTRIBUCION DE VOTOS - SENADO",
              "Votos por partido en cada municipio | Preconteo Congreso 2026", ncols)

    headers = ["Municipio", "Partido", "Votos", "% del Municipio", "Posicion", "Codigo Partido"]
    row = 4
    for ci, h in enumerate(headers, 1):
        ws.cell(row=row, column=ci, value=h)
    style_header_row(ws, row, ncols)
    row += 1

    c = conn.cursor()
    pie_data = {}

    for amb in AMBS:
        c.execute("SELECT SUM(votos) FROM votos WHERE corporacion='SE' AND amb=?", (amb,))
        total_muni = c.fetchone()[0] or 1

        c.execute("""
            SELECT m.nombre, COALESCE(p.nombre, 'Partido ' || v.codpar) as partido,
                   SUM(v.votos) as votos, v.codpar
            FROM votos v
            JOIN municipios m ON v.amb = m.amb
            LEFT JOIN partidos p ON v.codpar = p.codpar AND v.corporacion = p.corporacion
            WHERE v.corporacion = 'SE' AND v.amb = ?
            GROUP BY v.codpar
            ORDER BY votos DESC
        """, (amb,))

        results = c.fetchall()
        muni_name = results[0][0] if results else ""
        pie_data[muni_name] = []
        pos = 0
        for r in results:
            pos += 1
            ws.cell(row=row, column=1, value=r[0])
            ws.cell(row=row, column=2, value=r[1])
            ws.cell(row=row, column=3, value=r[2])
            ws.cell(row=row, column=4, value=r[2] / total_muni)
            ws.cell(row=row, column=4).number_format = '0.00%'
            ws.cell(row=row, column=5, value=pos)
            ws.cell(row=row, column=6, value=r[3])
            style_data_row(ws, row, ncols, alt=(row % 2 == 0))
            if pos <= 5:
                pie_data[muni_name].append((r[1], r[2]))
            row += 1

    # --- Grafica de torta para CADA municipio (2x2 grid) ---
    chart_positions = ["A", "E", "A", "E"]
    chart_row_offsets = [0, 0, 16, 16]

    base_row = row + 2
    for idx, (muni_name, pie_rows) in enumerate(pie_data.items()):
        if not pie_rows:
            continue
        pie_start = base_row + 34  # datos auxiliares lejos de las graficas
        pie_start += idx * 8

        ws.cell(row=pie_start, column=1, value="Partido")
        ws.cell(row=pie_start, column=2, value="Votos")
        style_header_row(ws, pie_start, 2)
        for i, (partido, votos) in enumerate(pie_rows):
            ws.cell(row=pie_start + 1 + i, column=1, value=partido)
            ws.cell(row=pie_start + 1 + i, column=2, value=votos)
            style_data_row(ws, pie_start + 1 + i, 2)

        pie = PieChart()
        pie.title = f"Top 5 Partidos Senado - {muni_name}"
        pie.style = 10
        pie.width = 16
        pie.height = 13
        labels = Reference(ws, min_col=1, min_row=pie_start + 1,
                           max_row=pie_start + len(pie_rows))
        data = Reference(ws, min_col=2, min_row=pie_start,
                         max_row=pie_start + len(pie_rows))
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        for ci, color in enumerate(CHART_COLORS[:len(pie_rows)]):
            pt = DataPoint(idx=ci)
            pt.graphicalProperties.solidFill = color
            pie.series[0].data_points.append(pt)

        col_letter = chart_positions[idx]
        row_offset = base_row + chart_row_offsets[idx]
        ws.add_chart(pie, f"{col_letter}{row_offset}")

    # --- Grafica de barras comparativa todos los municipios top 5 partidos ---
    comp_row = base_row + 34 + len(pie_data) * 8 + 2
    c2 = conn.cursor()
    c2.execute("""
        SELECT COALESCE(p.nombre, 'Partido ' || v.codpar) as partido,
               SUM(v.votos) as votos
        FROM votos v
        LEFT JOIN partidos p ON v.codpar = p.codpar AND v.corporacion = p.corporacion
        WHERE v.corporacion = 'SE' AND v.amb IN (?,?,?,?)
        GROUP BY v.codpar
        ORDER BY votos DESC
        LIMIT 8
    """, AMBS)

    ws.cell(row=comp_row, column=1, value="Partido").font = WHITE_FONT
    ws.cell(row=comp_row, column=1).fill = HEADER_BG
    ws.cell(row=comp_row, column=2, value="Total Votos SE").font = WHITE_FONT
    ws.cell(row=comp_row, column=2).fill = HEADER_BG
    top_partidos = c2.fetchall()
    for i, r in enumerate(top_partidos):
        ws.cell(row=comp_row + 1 + i, column=1, value=r[0])
        ws.cell(row=comp_row + 1 + i, column=2, value=r[1])
        style_data_row(ws, comp_row + 1 + i, 2)

    bar = BarChart()
    bar.type = "bar"
    bar.title = "Total votos Senado por partido (4 municipios)"
    bar.x_axis.title = "Votos"
    bar.style = 10
    bar.width = 24
    bar.height = 14
    bar.legend = None
    cats_b = Reference(ws, min_col=1, min_row=comp_row + 1,
                       max_row=comp_row + len(top_partidos))
    data_b = Reference(ws, min_col=2, min_row=comp_row,
                       max_row=comp_row + len(top_partidos))
    bar.add_data(data_b, titles_from_data=True)
    bar.set_categories(cats_b)
    bar.series[0].graphicalProperties.solidFill = "7B2D8B"
    ws.add_chart(bar, f"D{comp_row}")

    auto_width(ws, ncols)


def sheet_arrastre(wb, conn):
    ws = wb.create_sheet("Arrastre Verde")
    ws.sheet_properties.tabColor = "007C34"
    ncols = 9

    add_title(ws, "ARRASTRE ELECTORAL - ALIANZA VERDE (CA -> SE)",
              "Ratio votos_SE_Verde / votos_CA_Verde por zona | codpar CA=5, SE=57", ncols)

    headers = ["Municipio", "Zona", "Votos CA Verde", "Votos SE Verde",
               "Diferencia", "Variacion %", "Ratio SE/CA", "Tipo Arrastre", "Interpretacion"]
    row = 4
    for ci, h in enumerate(headers, 1):
        ws.cell(row=row, column=ci, value=h)
    style_header_row(ws, row, ncols)
    row += 1

    c = conn.cursor()
    c.execute("""
        SELECT m.nombre, ca.zona, ca.votos_ca, se.votos_se,
               ROUND(CAST(se.votos_se AS REAL) / ca.votos_ca, 3) as ratio
        FROM (
            SELECT amb, zona, SUM(votos) as votos_ca
            FROM votos WHERE corporacion='CA' AND codpar='5'
            GROUP BY amb, zona
        ) ca
        JOIN (
            SELECT amb, zona, SUM(votos) as votos_se
            FROM votos WHERE corporacion='SE' AND codpar='57'
            GROUP BY amb, zona
        ) se ON ca.amb = se.amb AND ca.zona = se.zona
        JOIN municipios m ON ca.amb = m.amb
        WHERE ca.votos_ca > 0
        ORDER BY ratio DESC
    """)

    chart_start = row
    all_rows = c.fetchall()
    for r in all_rows:
        diff = r[3] - r[2]
        var_pct = diff / r[2] if r[2] > 0 else 0
        tipo = "Positivo" if r[4] >= 1.0 else "Negativo"
        if r[4] >= 1.2:
            interp = "SE arrastra fuerte (+20%)"
        elif r[4] >= 1.0:
            interp = "Arrastre leve o neutro"
        elif r[4] >= 0.8:
            interp = "Perdida moderada"
        else:
            interp = "Perdida severa (>20%)"

        ws.cell(row=row, column=1, value=r[0])
        ws.cell(row=row, column=2, value=r[1])
        ws.cell(row=row, column=3, value=r[2])
        ws.cell(row=row, column=4, value=r[3])
        ws.cell(row=row, column=5, value=diff)
        ws.cell(row=row, column=6, value=var_pct)
        ws.cell(row=row, column=6).number_format = '0.0%'
        ws.cell(row=row, column=7, value=r[4])
        ws.cell(row=row, column=7).number_format = '0.000'
        ws.cell(row=row, column=8, value=tipo)
        ws.cell(row=row, column=9, value=interp)

        style_data_row(ws, row, ncols, alt=(row % 2 == 0))
        tipo_cell = ws.cell(row=row, column=8)
        if tipo == "Positivo":
            tipo_cell.font = Font(color="00C853", size=10, bold=True)
        else:
            tipo_cell.font = Font(color="FF5252", size=10, bold=True)
        row += 1
    chart_end = row - 1

    # Grafica 1: Barras del ratio por zona
    chart = BarChart()
    chart.type = "col"
    chart.title = "Ratio Arrastre Verde SE/CA por zona"
    chart.y_axis.title = "Ratio"
    chart.style = 10
    chart.width = 28
    chart.height = 14
    chart.legend = None
    cats = Reference(ws, min_col=2, min_row=chart_start, max_row=chart_end)
    data = Reference(ws, min_col=7, min_row=chart_start - 1, max_row=chart_end)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    # Colorear verde
    chart.series[0].graphicalProperties.solidFill = "007C34"
    ws.add_chart(chart, f"A{row + 1}")

    # Grafica 2: Barras agrupadas CA vs SE por zona
    chart2 = BarChart()
    chart2.type = "col"
    chart2.grouping = "clustered"
    chart2.title = "Votos Verde: Camara vs Senado por zona"
    chart2.y_axis.title = "Votos"
    chart2.style = 10
    chart2.width = 28
    chart2.height = 14
    cats2 = Reference(ws, min_col=2, min_row=chart_start, max_row=chart_end)
    d_ca = Reference(ws, min_col=3, min_row=chart_start - 1, max_row=chart_end)
    d_se = Reference(ws, min_col=4, min_row=chart_start - 1, max_row=chart_end)
    chart2.add_data(d_ca, titles_from_data=True)
    chart2.add_data(d_se, titles_from_data=True)
    chart2.set_categories(cats2)
    chart2.series[0].graphicalProperties.solidFill = "2196F3"
    chart2.series[1].graphicalProperties.solidFill = "007C34"
    ws.add_chart(chart2, f"A{row + 17}")

    # Grafica 3: Resumen por municipio (agregado)
    c.execute("""
        SELECT m.nombre,
            SUM(CASE WHEN v.corporacion='CA' AND v.codpar='5' THEN v.votos ELSE 0 END) as vca,
            SUM(CASE WHEN v.corporacion='SE' AND v.codpar='57' THEN v.votos ELSE 0 END) as vse
        FROM municipios m
        JOIN votos v ON m.amb = v.amb
        WHERE m.amb IN (?,?,?,?)
        GROUP BY m.amb ORDER BY m.nombre
    """, AMBS)

    summary_row = row + 34
    ws.cell(row=summary_row, column=1, value="Municipio").font = WHITE_FONT
    ws.cell(row=summary_row, column=1).fill = HEADER_BG
    ws.cell(row=summary_row, column=2, value="Votos CA Verde").font = WHITE_FONT
    ws.cell(row=summary_row, column=2).fill = HEADER_BG
    ws.cell(row=summary_row, column=3, value="Votos SE Verde").font = WHITE_FONT
    ws.cell(row=summary_row, column=3).fill = HEADER_BG
    ws.cell(row=summary_row, column=4, value="Ratio").font = WHITE_FONT
    ws.cell(row=summary_row, column=4).fill = HEADER_BG

    muni_data = c.fetchall()
    for i, r in enumerate(muni_data):
        ws.cell(row=summary_row + 1 + i, column=1, value=r[0])
        ws.cell(row=summary_row + 1 + i, column=2, value=r[1])
        ws.cell(row=summary_row + 1 + i, column=3, value=r[2])
        ws.cell(row=summary_row + 1 + i, column=4, value=r[2] / r[1] if r[1] > 0 else 0)
        ws.cell(row=summary_row + 1 + i, column=4).number_format = '0.000'
        style_data_row(ws, summary_row + 1 + i, 4)

    chart3 = BarChart()
    chart3.type = "col"
    chart3.grouping = "clustered"
    chart3.title = "Arrastre Verde por municipio (agregado)"
    chart3.y_axis.title = "Votos"
    chart3.style = 10
    chart3.width = 20
    chart3.height = 14
    cats3 = Reference(ws, min_col=1, min_row=summary_row + 1,
                      max_row=summary_row + len(muni_data))
    d_ca3 = Reference(ws, min_col=2, min_row=summary_row,
                      max_row=summary_row + len(muni_data))
    d_se3 = Reference(ws, min_col=3, min_row=summary_row,
                      max_row=summary_row + len(muni_data))
    chart3.add_data(d_ca3, titles_from_data=True)
    chart3.add_data(d_se3, titles_from_data=True)
    chart3.set_categories(cats3)
    chart3.series[0].graphicalProperties.solidFill = "2196F3"
    chart3.series[1].graphicalProperties.solidFill = "007C34"
    ws.add_chart(chart3, f"F{summary_row}")

    auto_width(ws, ncols)


def sheet_dominancia(wb, conn):
    ws = wb.create_sheet("Dominancia >60%")
    ws.sheet_properties.tabColor = "FF5252"
    ncols = 9

    add_title(ws, "DOMINANCIA EXTREMA DE CANDIDATOS",
              "Zonas donde un candidato concentra >60% de los votos de su partido", ncols)

    headers = ["Municipio", "Zona", "Corporacion", "Partido", "Candidato",
               "Votos Candidato", "Votos Partido Zona", "Dominancia %", "Nivel"]
    row = 4
    for ci, h in enumerate(headers, 1):
        ws.cell(row=row, column=ci, value=h)
    style_header_row(ws, row, ncols)
    row += 1

    c = conn.cursor()
    c.execute("""
        SELECT m.nombre, v.zona, v.corporacion,
               COALESCE(p.nombre, 'Partido ' || v.codpar) as partido,
               ca.nomcan || ' ' || ca.apecan as candidato,
               v.votos as votos_cand,
               partido_zona.votos_partido,
               ROUND(CAST(v.votos AS REAL) / partido_zona.votos_partido * 100, 1) as pct
        FROM votos v
        JOIN municipios m ON v.amb = m.amb
        JOIN candidatos ca ON v.codcan = ca.codcan AND v.codpar = ca.codpar
            AND v.corporacion = ca.corporacion AND v.amb = ca.amb
        LEFT JOIN partidos p ON v.codpar = p.codpar AND v.corporacion = p.corporacion
        JOIN (
            SELECT corporacion, amb, zona, codpar, SUM(votos) as votos_partido
            FROM votos GROUP BY corporacion, amb, zona, codpar
        ) partido_zona ON v.corporacion = partido_zona.corporacion
            AND v.amb = partido_zona.amb AND v.zona = partido_zona.zona
            AND v.codpar = partido_zona.codpar
        WHERE v.codcan != '0' AND partido_zona.votos_partido > 0
            AND CAST(v.votos AS REAL) / partido_zona.votos_partido > 0.60
        ORDER BY pct DESC
    """)

    all_rows = c.fetchall()
    counts = {"Absoluta": 0, "Alta": 0, "Moderada": 0}

    for r in all_rows:
        pct = r[7]
        if pct >= 90:
            nivel = "Absoluta"
        elif pct >= 75:
            nivel = "Alta"
        else:
            nivel = "Moderada"
        counts[nivel] += 1

        ws.cell(row=row, column=1, value=r[0])
        ws.cell(row=row, column=2, value=r[1])
        ws.cell(row=row, column=3, value=r[2])
        ws.cell(row=row, column=4, value=r[3])
        ws.cell(row=row, column=5, value=r[4])
        ws.cell(row=row, column=6, value=r[5])
        ws.cell(row=row, column=7, value=r[6])
        ws.cell(row=row, column=8, value=pct / 100)
        ws.cell(row=row, column=8).number_format = '0.0%'
        ws.cell(row=row, column=9, value=nivel)
        style_data_row(ws, row, ncols, alt=(row % 2 == 0))

        nivel_cell = ws.cell(row=row, column=9)
        if nivel == "Absoluta":
            nivel_cell.font = Font(color="FF5252", size=10, bold=True)
        elif nivel == "Alta":
            nivel_cell.font = Font(color="E07B00", size=10, bold=True)
        else:
            nivel_cell.font = Font(color="2196F3", size=10)
        row += 1

    # --- Grafica 1: Torta de niveles de dominancia ---
    pie_row = row + 2
    ws.cell(row=pie_row, column=1, value="Nivel").font = WHITE_FONT
    ws.cell(row=pie_row, column=1).fill = HEADER_BG
    ws.cell(row=pie_row, column=2, value="Cantidad").font = WHITE_FONT
    ws.cell(row=pie_row, column=2).fill = HEADER_BG
    for i, (nivel, cnt) in enumerate(counts.items()):
        ws.cell(row=pie_row + 1 + i, column=1, value=nivel)
        ws.cell(row=pie_row + 1 + i, column=2, value=cnt)
        style_data_row(ws, pie_row + 1 + i, 2)

    pie = PieChart()
    pie.title = "Distribucion por nivel de dominancia"
    pie.style = 10
    pie.width = 16
    pie.height = 13
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showVal = True
    labels = Reference(ws, min_col=1, min_row=pie_row + 1, max_row=pie_row + 3)
    data = Reference(ws, min_col=2, min_row=pie_row, max_row=pie_row + 3)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    for ci, color in enumerate(["FF5252", "E07B00", "2196F3"]):
        pt = DataPoint(idx=ci)
        pt.graphicalProperties.solidFill = color
        pie.series[0].data_points.append(pt)
    ws.add_chart(pie, f"D{pie_row}")

    # --- Grafica 2: Top 10 candidatos dominantes (barras) ---
    bar_row = pie_row + 16
    ws.cell(row=bar_row, column=1, value="Candidato (zona)").font = WHITE_FONT
    ws.cell(row=bar_row, column=1).fill = HEADER_BG
    ws.cell(row=bar_row, column=2, value="Dominancia %").font = WHITE_FONT
    ws.cell(row=bar_row, column=2).fill = HEADER_BG

    top10_dom = all_rows[:10]
    for i, r in enumerate(top10_dom):
        label = f"{r[4]} (Z{r[1][-2:]})"
        ws.cell(row=bar_row + 1 + i, column=1, value=label)
        ws.cell(row=bar_row + 1 + i, column=2, value=r[7])
        ws.cell(row=bar_row + 1 + i, column=2).number_format = '0.0'
        style_data_row(ws, bar_row + 1 + i, 2)

    bar = BarChart()
    bar.type = "bar"
    bar.title = "Top 10 candidatos con mayor dominancia"
    bar.x_axis.title = "% Dominancia"
    bar.style = 10
    bar.width = 28
    bar.height = 14
    bar.legend = None
    cats_b = Reference(ws, min_col=1, min_row=bar_row + 1,
                       max_row=bar_row + len(top10_dom))
    data_b = Reference(ws, min_col=2, min_row=bar_row,
                       max_row=bar_row + len(top10_dom))
    bar.add_data(data_b, titles_from_data=True)
    bar.set_categories(cats_b)
    bar.series[0].graphicalProperties.solidFill = "FF5252"
    ws.add_chart(bar, f"D{bar_row}")

    # --- Grafica 3: Conteo de zonas dominantes por municipio ---
    c2 = conn.cursor()
    c2.execute("""
        SELECT m.nombre, COUNT(*) as zonas_dominantes
        FROM votos v
        JOIN municipios m ON v.amb = m.amb
        JOIN (
            SELECT corporacion, amb, zona, codpar, SUM(votos) as votos_partido
            FROM votos GROUP BY corporacion, amb, zona, codpar
        ) pz ON v.corporacion = pz.corporacion AND v.amb = pz.amb
            AND v.zona = pz.zona AND v.codpar = pz.codpar
        WHERE v.codcan != '0' AND pz.votos_partido > 0
            AND CAST(v.votos AS REAL) / pz.votos_partido > 0.60
            AND v.amb IN (?,?,?,?)
        GROUP BY m.nombre
        ORDER BY zonas_dominantes DESC
    """, AMBS)

    muni_dom_row = bar_row + len(top10_dom) + 3
    ws.cell(row=muni_dom_row, column=1, value="Municipio").font = WHITE_FONT
    ws.cell(row=muni_dom_row, column=1).fill = HEADER_BG
    ws.cell(row=muni_dom_row, column=2, value="Zonas con dominancia").font = WHITE_FONT
    ws.cell(row=muni_dom_row, column=2).fill = HEADER_BG
    muni_dom_data = c2.fetchall()
    for i, r in enumerate(muni_dom_data):
        ws.cell(row=muni_dom_row + 1 + i, column=1, value=r[0])
        ws.cell(row=muni_dom_row + 1 + i, column=2, value=r[1])
        style_data_row(ws, muni_dom_row + 1 + i, 2)

    bar3 = BarChart()
    bar3.type = "col"
    bar3.title = "Zonas con dominancia >60% por municipio"
    bar3.y_axis.title = "Cantidad de zonas"
    bar3.style = 10
    bar3.width = 18
    bar3.height = 12
    bar3.legend = None
    cats3 = Reference(ws, min_col=1, min_row=muni_dom_row + 1,
                      max_row=muni_dom_row + len(muni_dom_data))
    data3 = Reference(ws, min_col=2, min_row=muni_dom_row,
                      max_row=muni_dom_row + len(muni_dom_data))
    bar3.add_data(data3, titles_from_data=True)
    bar3.set_categories(cats3)
    bar3.series[0].graphicalProperties.solidFill = "E07B00"
    ws.add_chart(bar3, f"D{muni_dom_row}")

    auto_width(ws, ncols)


def sheet_atribucion(wb, conn):
    ws = wb.create_sheet("Atribucion SE")
    ws.sheet_properties.tabColor = "1E477D"
    ncols = 8

    add_title(ws, "ATRIBUCION DETERMINISTICA CA -> SE",
              "Formula: A = (votos_cand / votos_partido_CA) x votos_SE_partido | Top 20", ncols)

    headers = ["Municipio", "Candidato", "Partido CA", "Votos Candidato",
               "Votos Partido CA", "Votos Partido SE", "Atribucion SE", "% Atribucion"]
    row = 4
    for ci, h in enumerate(headers, 1):
        ws.cell(row=row, column=ci, value=h)
    style_header_row(ws, row, ncols)
    row += 1

    c = conn.cursor()
    c.execute("""
        SELECT m.nombre,
               ca_cand.nomcan || ' ' || ca_cand.apecan as candidato,
               COALESCE(p.nombre, 'Partido ' || v.codpar) as partido,
               SUM(v.votos) as votos_cand,
               ca_total.votos_partido_ca,
               se_total.votos_partido_se,
               ROUND(CAST(SUM(v.votos) AS REAL) / ca_total.votos_partido_ca * se_total.votos_partido_se, 1) as atribucion
        FROM votos v
        JOIN municipios m ON v.amb = m.amb
        JOIN candidatos ca_cand ON v.codcan = ca_cand.codcan AND v.codpar = ca_cand.codpar
            AND v.corporacion = ca_cand.corporacion AND v.amb = ca_cand.amb
        LEFT JOIN partidos p ON v.codpar = p.codpar AND v.corporacion = p.corporacion
        JOIN (
            SELECT amb, codpar, SUM(votos) as votos_partido_ca
            FROM votos WHERE corporacion='CA'
            GROUP BY amb, codpar
        ) ca_total ON v.amb = ca_total.amb AND v.codpar = ca_total.codpar
        JOIN (
            SELECT amb, codpar, SUM(votos) as votos_partido_se
            FROM votos WHERE corporacion='SE'
            GROUP BY amb, codpar
        ) se_total ON v.amb = se_total.amb
            AND CASE v.codpar
                WHEN '5' THEN '57'
                WHEN '87' THEN '92'
                WHEN '10' THEN '10'
                WHEN '2' THEN '2'
                ELSE NULL
            END = se_total.codpar
        WHERE v.corporacion = 'CA' AND v.codcan != '0'
            AND ca_total.votos_partido_ca > 0
        GROUP BY v.amb, v.codpar, v.codcan
        ORDER BY atribucion DESC
        LIMIT 20
    """)

    total_atrib = 0
    results = c.fetchall()
    for r in results:
        total_atrib += r[6]

    chart_start = row
    for r in results:
        ws.cell(row=row, column=1, value=r[0])
        ws.cell(row=row, column=2, value=r[1])
        ws.cell(row=row, column=3, value=r[2])
        ws.cell(row=row, column=4, value=r[3])
        ws.cell(row=row, column=5, value=r[4])
        ws.cell(row=row, column=6, value=r[5])
        ws.cell(row=row, column=7, value=r[6])
        ws.cell(row=row, column=8, value=r[6] / total_atrib if total_atrib > 0 else 0)
        ws.cell(row=row, column=8).number_format = '0.00%'
        style_data_row(ws, row, ncols, alt=(row % 2 == 0))
        row += 1
    chart_end = row - 1

    # --- Grafica 1: Barras horizontales top 20 atribucion ---
    chart = BarChart()
    chart.type = "bar"
    chart.title = "Top 20 candidatos por atribucion al Senado"
    chart.x_axis.title = "Atribucion (votos equivalentes SE)"
    chart.style = 10
    chart.width = 30
    chart.height = 18
    chart.legend = None
    cats = Reference(ws, min_col=2, min_row=chart_start, max_row=chart_end)
    data = Reference(ws, min_col=7, min_row=chart_start - 1, max_row=chart_end)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = "1E477D"
    ws.add_chart(chart, f"A{row + 1}")

    # --- Grafica 2: Torta por partido (atribucion agrupada) ---
    from collections import defaultdict
    partido_atrib = defaultdict(float)
    for r in results:
        partido_atrib[r[2]] += r[6]

    pie_row = row + 20
    ws.cell(row=pie_row, column=1, value="Partido").font = WHITE_FONT
    ws.cell(row=pie_row, column=1).fill = HEADER_BG
    ws.cell(row=pie_row, column=2, value="Atribucion Total").font = WHITE_FONT
    ws.cell(row=pie_row, column=2).fill = HEADER_BG

    sorted_partidos = sorted(partido_atrib.items(), key=lambda x: x[1], reverse=True)
    for i, (partido, atrib) in enumerate(sorted_partidos):
        ws.cell(row=pie_row + 1 + i, column=1, value=partido)
        ws.cell(row=pie_row + 1 + i, column=2, value=atrib)
        ws.cell(row=pie_row + 1 + i, column=2).number_format = '#,##0.0'
        style_data_row(ws, pie_row + 1 + i, 2)

    pie = PieChart()
    pie.title = "Atribucion SE por partido"
    pie.style = 10
    pie.width = 18
    pie.height = 14
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showCatName = True
    labels = Reference(ws, min_col=1, min_row=pie_row + 1,
                       max_row=pie_row + len(sorted_partidos))
    data_p = Reference(ws, min_col=2, min_row=pie_row,
                       max_row=pie_row + len(sorted_partidos))
    pie.add_data(data_p, titles_from_data=True)
    pie.set_categories(labels)
    for ci, color in enumerate(CHART_COLORS[:len(sorted_partidos)]):
        pt = DataPoint(idx=ci)
        pt.graphicalProperties.solidFill = color
        pie.series[0].data_points.append(pt)
    ws.add_chart(pie, f"D{pie_row}")

    # --- Grafica 3: Comparativa votos directos CA vs atribucion SE ---
    comp_row = pie_row + len(sorted_partidos) + 3
    ws.cell(row=comp_row, column=1, value="Candidato").font = WHITE_FONT
    ws.cell(row=comp_row, column=1).fill = HEADER_BG
    ws.cell(row=comp_row, column=2, value="Votos CA directos").font = WHITE_FONT
    ws.cell(row=comp_row, column=2).fill = HEADER_BG
    ws.cell(row=comp_row, column=3, value="Atribucion SE").font = WHITE_FONT
    ws.cell(row=comp_row, column=3).fill = HEADER_BG

    top10 = results[:10]
    for i, r in enumerate(top10):
        ws.cell(row=comp_row + 1 + i, column=1, value=r[1])
        ws.cell(row=comp_row + 1 + i, column=2, value=r[3])
        ws.cell(row=comp_row + 1 + i, column=3, value=r[6])
        style_data_row(ws, comp_row + 1 + i, 3)

    chart3 = BarChart()
    chart3.type = "bar"
    chart3.grouping = "clustered"
    chart3.title = "Top 10: Votos CA directos vs Atribucion SE"
    chart3.x_axis.title = "Votos"
    chart3.style = 10
    chart3.width = 28
    chart3.height = 16
    cats3 = Reference(ws, min_col=1, min_row=comp_row + 1, max_row=comp_row + len(top10))
    d_ca = Reference(ws, min_col=2, min_row=comp_row, max_row=comp_row + len(top10))
    d_se = Reference(ws, min_col=3, min_row=comp_row, max_row=comp_row + len(top10))
    chart3.add_data(d_ca, titles_from_data=True)
    chart3.add_data(d_se, titles_from_data=True)
    chart3.set_categories(cats3)
    chart3.series[0].graphicalProperties.solidFill = "2196F3"
    chart3.series[1].graphicalProperties.solidFill = "1E477D"
    ws.add_chart(chart3, f"A{comp_row}")

    auto_width(ws, ncols)


def main():
    if not os.path.exists(DB_PATH):
        print(f"ERROR: No se encontro {DB_PATH}")
        return

    conn = get_conn()
    wb = Workbook()

    print("Generando reporte Excel...")
    print("  [1/6] Resumen general + 3 graficas (barras, torta, apiladas)")
    sheet_resumen(wb, conn)
    print("  [2/6] Top candidatos Camara + 2 graficas (top 10 global, podio por municipio)")
    sheet_candidatos_ca(wb, conn)
    print("  [3/6] Partidos Senado + 5 graficas (4 tortas + barras comparativas)")
    sheet_partidos_se(wb, conn)
    print("  [4/6] Arrastre Verde CA->SE + 3 graficas (ratio, CA vs SE, resumen municipio)")
    sheet_arrastre(wb, conn)
    print("  [5/6] Dominancia >60% + 3 graficas (torta niveles, top 10, por municipio)")
    sheet_dominancia(wb, conn)
    print("  [6/6] Atribucion SE + 3 graficas (top 20, torta partidos, CA vs atribucion)")
    sheet_atribucion(wb, conn)

    conn.close()
    wb.save(OUTPUT)
    print(f"\n  Guardado: {OUTPUT}")
    print(f"  {len(wb.sheetnames)} hojas, ~19 graficas: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
